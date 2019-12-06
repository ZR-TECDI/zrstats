from openpyxl import load_workbook
import datetime


workbook = load_workbook(filename="Asistencia_ZR_2019.xlsx")
print(workbook.sheetnames)

# miembros = []
# for sheet in workbook.worksheets:
#     print(sheet)
#     for row in sheet.iter_rows(min_row=3, max_row=100):
#         if row[2].value is not None:
#             # print(str(row[2].value) + "" + str(row[3].value))
#             pass


misiones = []
year = 2019
mes = 0
lista_misiones = []
for sheet in workbook.worksheets:
    if sheet.title == "Enero":
        mes = 1
    if sheet.title == "Febrero":
        mes = 2
    if sheet.title == "Marzo":
        mes = 3
    if sheet.title == "Abril":
        mes = 4
    if sheet.title == "Mayo":
        mes = 5
    if sheet.title == "Junio":
        mes = 6
    if sheet.title == "Julio":
        mes = 7
    if sheet.title == "Agosto":
        mes = 8
    if sheet.title == "Septiembre":
        mes = 9
    if sheet.title == "Octubre":
        mes = 10
    if sheet.title == "Noviembre":
        mes = 11
    if sheet.title == "Diciembre":
        mes = 12


    # RECORRO LOS DIAS DE CADA MES (columnas entre "K" y la primera columna con signo "%")
    for col in sheet["K:Y"]: # TODO AJUSTAR ACÁ EL RANGO DONDE ESTAN COLUMNAS DE LOS DIAS
        if col[1].value != "%" and col[1].value is not None:
            if isinstance(col[1].value, float):
                day = int(col[1].value)
            else:
                day = col[1].value.day
            dt = datetime.date(year, mes, day)
            lista_misiones.append(dt)

print(lista_misiones)
"""Agrega los datos iniciales de forma automática"""

from stats.models import *
from collections import defaultdict
import os
import random
import datetime
from urllib.request import urlopen
from django.core.files import File
from django.core.files.temp import NamedTemporaryFile
from zrstats import settings
from django.contrib.staticfiles.templatetags.staticfiles import static


def borra_todo():
    lista = [Miembro, Rango, Unidad, Rol, Nacionalidad, Clase, User, Asistencia, Mision, Campana]
    # lista = [Miembro, User, Asistencia, Mision, Campana]
    for item in lista:
        item.objects.all().delete()


def crear_unidades():
    unidades_dict = {
        'altm': ['Alto Mando', 'ALTM', 'http://www.zrarmy.com/images/Estrella_Dorada2.png'],
        'infm': ['Inf. de Marines', 'IMZR', 'http://www.zrarmy.com/images/INFANTERIA.png'],
        'paraca': ['Paracaidistas', 'PAR', 'http://www.zrarmy.com/images/PARACAIDISTAS.png'],
        'echo': ['Espectro', 'ECHO', 'http://www.zrarmy.com/images/SPECTRO.png'],
        'cab': ['Caballería', 'CAB', 'http://www.zrarmy.com/images/BLINDADOS.png'],
        'fazr': ['Fuerza Aérea', 'FAZR', 'http://www.zrarmy.com/images/FAZR.png'],
    }

    print('Creando Unidades...')

    for var, array in unidades_dict.items():
        var = Unidad.objects.create()
        var.nombre = array[0]
        var.abreviatura = array[1]
        img_temp = NamedTemporaryFile()
        img_temp.write(urlopen(array[2]).read())
        img_temp.flush()
        var.imagen.save(array[1] + ".png", File(img_temp))
        var.save()


def crear_rango():
    rangos_dict = {
        'may': ['Mayor', 'May', 0],
        'cap': ['Capitán', 'Cap', 1],
        'tte': ['Teniente', 'Tte', 2],
        'alf': ['Alférez', 'Alf', 3],
        'sgtm': ['Sargento Mayor', 'SgtM', 4],
        'sgt1': ['Sargento 1º', 'Sgt1', 5],
        'sgt': ['Sargento', 'Sgt', 6],
        'cbo1': ['Cabo 1º', 'Cbo1', 7],
        'cbo': ["Cabo", 'Cbo', 8],
        'dis': ['Distinguido', 'Dis', 9],
        'inf': ['Infante', 'Inf', 10],
        'rct': ['Recluta', 'Rct', 11],
        'asp': ['Aspirante', 'Asp', 12]
    }

    print('Creando rangos...')

    for var, array in rangos_dict.items():
        var = Rango.objects.create()
        var.rango = array[0]
        var.abreviatura = array[1]
        var.orden = array[2]
        var.save()


def crear_rol():
    roles_dict = {
        'nada': ['--', '--'],
        'cmd': ['Comandante Unidad', 'CMD'],
        'ce': ['Comandante Escuadra', 'CE'],
        'lea': ['Líder Equipo A', 'LEA'],
        'leb': ['Líder Equipo B', 'LEB'],
        'a': ['ALPHA', 'A'],
        'b': ['BRAVO', 'B'],
        'cmda1': ['Com. Ariete 1', 'CMDA1'],
        'cmda2': ['Com. Ariete 2', 'CMDA2'],
        'cmda3': ['Com. Ariete 3', 'CMDA3'],
        'cmda4': ['Com. Ariete 4', 'CMDA4']
    }

    print('Creando roles...')

    for var, array in roles_dict.items():
        var = Rol.objects.create()
        var.rol = array[0]
        var.abreviatura = array[1]
        var.save()


def crear_clase():
    clases_dict = {
        'nada': ['--', '--'],
        'fl': ['Fusilero', 'FL'],
        'mg': ['Ametrallador', 'MG'],
        'at': ['Antitanque', 'AT'],
        'gl': ['Granadero', 'GL'],
        'mc': ['Médico de combate', 'MC'],
        'od': ['Operador de dron', 'OD'],
        'ro': ['Radio operador', 'RO'],
        'ts': ['Tirador Selecto', 'TS'],
        'te': ['Tirador de escuadra', 'TE']
    }

    print('Creando clases...')

    for var, array in clases_dict.items():
        var = Clase.objects.create()
        var.clase = array[0]
        var.abreviatura = array[1]
        var.save()


def crear_naciones():
    naciones_dict = {
        'AF': ['Afghanistan', 'AF'],
        'AL': ['Albania', 'AL'],
        'DZ': ['Algeria', 'DZ'],
        'AS': ['AmericanSamoa', 'AS'],
        'AD': ['Andorra', 'AD'],
        'AO': ['Angola', 'AO'],
        'AI': ['Anguilla', 'AI'],
        'AQ': ['Antarctica', 'AQ'],
        'AG': ['AntiguaandBarbuda', 'AG'],
        'AR': ['Argentina', 'AR'],
        'AM': ['Armenia', 'AM'],
        'AW': ['Aruba', 'AW'],
        'AU': ['Australia', 'AU'],
        'AT': ['Austria', 'AT'],
        'AZ': ['Azerbaijan', 'AZ'],
        'BS': ['Bahamas', 'BS'],
        'BH': ['Bahrain', 'BH'],
        'BD': ['Bangladesh', 'BD'],
        'BB': ['Barbados', 'BB'],
        'BY': ['Belarus', 'BY'],
        'BE': ['Belgium', 'BE'],
        'BZ': ['Belize', 'BZ'],
        'BJ': ['Benin', 'BJ'],
        'BM': ['Bermuda', 'BM'],
        'BT': ['Bhutan', 'BT'],
        'BO': ['Bolivia', 'BO'],
        'BA': ['BosniaandHerzegovina', 'BA'],
        'BW': ['Botswana', 'BW'],
        'BV': ['BouvetIsland', 'BV'],
        'BR': ['Brazil', 'BR'],
        'IO': ['BritishIndianOceanTerritory', 'IO'],
        'BN': ['Brunei', 'BN'],
        'BG': ['Bulgaria', 'BG'],
        'BF': ['BurkinaFaso', 'BF'],
        'BI': ['Burundi', 'BI'],
        'KH': ['Cambodia', 'KH'],
        'CM': ['Cameroon', 'CM'],
        'CA': ['Canada', 'CA'],
        'CV': ['CapeVerde', 'CV'],
        'KY': ['CaymanIslands', 'KY'],
        'CF': ['CentralAfricanRepublic', 'CF'],
        'TD': ['Chad', 'TD'],
        'CL': ['Chile', 'CL'],
        'CN': ['China', 'CN'],
        'CX': ['ChristmasIsland', 'CX'],
        'CC': ['Cocos(Keeling)Islands', 'CC'],
        'CO': ['Colombia', 'CO'],
        'KM': ['Comoros', 'KM'],
        'CG': ['Congo', 'CG'],
        'CD': ['Congo,theDemocraticRepublicofthe', 'CD'],
        'CK': ['CookIslands', 'CK'],
        'CR': ['CostaRica', 'CR'],
        'CI': ['IvoryCoast', 'CI'],
        'HR': ['Croatia', 'HR'],
        'CU': ['Cuba', 'CU'],
        'CY': ['Cyprus', 'CY'],
        'CZ': ['CzechRepublic', 'CZ'],
        'DK': ['Denmark', 'DK'],
        'DJ': ['Djibouti', 'DJ'],
        'DM': ['Dominica', 'DM'],
        'DO': ['DominicanRepublic', 'DO'],
        'EC': ['Ecuador', 'EC'],
        'EG': ['Egypt', 'EG'],
        'SV': ['ElSalvador', 'SV'],
        'GQ': ['EquatorialGuinea', 'GQ'],
        'ER': ['Eritrea', 'ER'],
        'EE': ['Estonia', 'EE'],
        'ET': ['Ethiopia', 'ET'],
        'FK': ['FalklandIslands(Malvinas)', 'FK'],
        'FO': ['FaroeIslands', 'FO'],
        'FJ': ['Fiji', 'FJ'],
        'FI': ['Finland', 'FI'],
        'FR': ['France', 'FR'],
        'GF': ['FrenchGuiana', 'GF'],
        'PF': ['FrenchPolynesia', 'PF'],
        'TF': ['FrenchSouthernTerritories', 'TF'],
        'GA': ['Gabon', 'GA'],
        'GM': ['Gambia', 'GM'],
        'GE': ['Georgia', 'GE'],
        'DE': ['Germany', 'DE'],
        'GH': ['Ghana', 'GH'],
        'GI': ['Gibraltar', 'GI'],
        'GR': ['Greece', 'GR'],
        'GL': ['Greenland', 'GL'],
        'GD': ['Grenada', 'GD'],
        'GP': ['Guadeloupe', 'GP'],
        'GU': ['Guam', 'GU'],
        'GT': ['Guatemala', 'GT'],
        'GG': ['Guernsey', 'GG'],
        'GN': ['Guinea', 'GN'],
        'GW': ['Guinea-Bissau', 'GW'],
        'GY': ['Guyana', 'GY'],
        'HT': ['Haiti', 'HT'],
        'HM': ['HeardIslandandMcDonaldIslands', 'HM'],
        'VA': ['HolySee(VaticanCityState)', 'VA'],
        'HN': ['Honduras', 'HN'],
        'HK': ['HongKong', 'HK'],
        'HU': ['Hungary', 'HU'],
        'IS': ['Iceland', 'IS'],
        'IN': ['India', 'IN'],
        'ID': ['Indonesia', 'ID'],
        'IR': ['Iran,IslamicRepublicof', 'IR'],
        'IQ': ['Iraq', 'IQ'],
        'IE': ['Ireland', 'IE'],
        'IM': ['IsleofMan', 'IM'],
        'IL': ['Israel', 'IL'],
        'IT': ['Italy', 'IT'],
        'JM': ['Jamaica', 'JM'],
        'JP': ['Japan', 'JP'],
        'JE': ['Jersey', 'JE'],
        'JO': ['Jordan', 'JO'],
        'KZ': ['Kazakhstan', 'KZ'],
        'KE': ['Kenya', 'KE'],
        'KI': ['Kiribati', 'KI'],
        'KP': ['Korea,DemocraticPeoplesRepublicof', 'KP'],
        'KR': ['SouthKorea', 'KR'],
        'KW': ['Kuwait', 'KW'],
        'KG': ['Kyrgyzstan', 'KG'],
        'LA': ['LaoPeoplesDemocraticRepublic', 'LA'],
        'LV': ['Latvia', 'LV'],
        'LB': ['Lebanon', 'LB'],
        'LS': ['Lesotho', 'LS'],
        'LR': ['Liberia', 'LR'],
        'LY': ['Libya', 'LY'],
        'LI': ['Liechtenstein', 'LI'],
        'LT': ['Lithuania', 'LT'],
        'LU': ['Luxembourg', 'LU'],
        'MO': ['Macao', 'MO'],
        'MK': ['Macedonia,theformerYugoslavRepublicof', 'MK'],
        'MG': ['Madagascar', 'MG'],
        'MW': ['Malawi', 'MW'],
        'MY': ['Malaysia', 'MY'],
        'MV': ['Maldives', 'MV'],
        'ML': ['Mali', 'ML'],
        'MT': ['Malta', 'MT'],
        'MH': ['MarshallIslands', 'MH'],
        'MQ': ['Martinique', 'MQ'],
        'MR': ['Mauritania', 'MR'],
        'MU': ['Mauritius', 'MU'],
        'YT': ['Mayotte', 'YT'],
        'MX': ['Mexico', 'MX'],
        'FM': ['Micronesia,FederatedStatesof', 'FM'],
        'MD': ['Moldova,Republicof', 'MD'],
        'MC': ['Monaco', 'MC'],
        'MN': ['Mongolia', 'MN'],
        'ME': ['Montenegro', 'ME'],
        'MS': ['Montserrat', 'MS'],
        'MA': ['Morocco', 'MA'],
        'MZ': ['Mozambique', 'MZ'],
        'MM': ['Birmania', 'MM'],
        'NA': ['Namibia', 'NA'],
        'NR': ['Nauru', 'NR'],
        'NP': ['Nepal', 'NP'],
        'NL': ['Netherlands', 'NL'],
        'AN': ['NetherlandsAntilles', 'AN'],
        'NC': ['NewCaledonia', 'NC'],
        'NZ': ['NewZealand', 'NZ'],
        'NI': ['Nicaragua', 'NI'],
        'NE': ['Niger', 'NE'],
        'NG': ['Nigeria', 'NG'],
        'NU': ['Niue', 'NU'],
        'NF': ['NorfolkIsland', 'NF'],
        'MP': ['NorthernMarianaIslands', 'MP'],
        'NO': ['Norway', 'NO'],
        'OM': ['Oman', 'OM'],
        'PK': ['Pakistan', 'PK'],
        'PW': ['Palau', 'PW'],
        'PS': ['PalestinianTerritory,Occupied', 'PS'],
        'PA': ['Panama', 'PA'],
        'PG': ['PapuaNewGuinea', 'PG'],
        'PY': ['Paraguay', 'PY'],
        'PE': ['Peru', 'PE'],
        'PH': ['Philippines', 'PH'],
        'PN': ['Pitcairn', 'PN'],
        'PL': ['Poland', 'PL'],
        'PT': ['Portugal', 'PT'],
        'PR': ['PuertoRico', 'PR'],
        'QA': ['Qatar', 'QA'],
        'RE': ['Réunion', 'RE'],
        'RO': ['Romania', 'RO'],
        'RU': ['Russia', 'RU'],
        'RW': ['Rwanda', 'RW'],
        'SH': ['SaintHelena', 'SH'],
        'KN': ['SaintKittsandNevis', 'KN'],
        'LC': ['SaintLucia', 'LC'],
        'PM': ['SaintPierreandMiquelon', 'PM'],
        'VC': ['St.VincentandtheGrenadines', 'VC'],
        'WS': ['Samoa', 'WS'],
        'SM': ['SanMarino', 'SM'],
        'ST': ['SaoTomeandPrincipe', 'ST'],
        'SA': ['SaudiArabia', 'SA'],
        'SN': ['Senegal', 'SN'],
        'RS': ['Serbia', 'RS'],
        'SC': ['Seychelles', 'SC'],
        'SL': ['SierraLeone', 'SL'],
        'SG': ['Singapore', 'SG'],
        'SK': ['Slovakia', 'SK'],
        'SI': ['Slovenia', 'SI'],
        'SB': ['SolomonIslands', 'SB'],
        'SO': ['Somalia', 'SO'],
        'ZA': ['SouthAfrica', 'ZA'],
        'GS': ['SouthGeorgiaandtheSouthSandwichIslands', 'GS'],
        'ES': ['Spain', 'ES'],
        'LK': ['SriLanka', 'LK'],
        'SD': ['Sudan', 'SD'],
        'SR': ['Suriname', 'SR'],
        'SJ': ['SvalbardandJanMayen', 'SJ'],
        'SZ': ['Swaziland', 'SZ'],
        'SE': ['Sweden', 'SE'],
        'CH': ['Switzerland', 'CH'],
        'SY': ['SyrianArabRepublic', 'SY'],
        'TW': ['Taiwan', 'TW'],
        'TJ': ['Tajikistan', 'TJ'],
        'TZ': ['Tanzania,UnitedRepublicof', 'TZ'],
        'TH': ['Thailand', 'TH'],
        'TL': ['Timor-Leste', 'TL'],
        'TG': ['Togo', 'TG'],
        'TK': ['Tokelau', 'TK'],
        'TO': ['Tonga', 'TO'],
        'TT': ['TrinidadandTobago', 'TT'],
        'TN': ['Tunisia', 'TN'],
        'TR': ['Turkey', 'TR'],
        'TM': ['Turkmenistan', 'TM'],
        'TC': ['TurksandCaicosIslands', 'TC'],
        'TV': ['Tuvalu', 'TV'],
        'UG': ['Uganda', 'UG'],
        'UA': ['Ukraine', 'UA'],
        'AE': ['UnitedArabEmirates', 'AE'],
        'GB': ['UnitedKingdom', 'GB'],
        'US': ['UnitedStates', 'US'],
        'UM': ['UnitedStatesMinorOutlyingIslands', 'UM'],
        'UY': ['Uruguay', 'UY'],
        'UZ': ['Uzbekistan', 'UZ'],
        'VU': ['Vanuatu', 'VU'],
        'VE': ['Venezuela', 'VE'],
        'VN': ['Vietnam', 'VN'],
        'VG': ['VirginIslands,British', 'VG'],
        'VI': ['VirginIslands,U.S.', 'VI'],
        'WF': ['WallisandFutuna', 'WF'],
        'EH': ['WesternSahara', 'EH'],
        'YE': ['Yemen', 'YE'],
        'ZM': ['Zambia', 'ZM'],
        'ZW': ['Zimbabwe', 'ZW'],
    }

    print('Creando naciones...')

    for var, array in naciones_dict.items():
        var = Nacionalidad.objects.create()
        var.pais = array[0]
        var.abreviatura = array[1]
        var.save()


def agregar_miembros():
    # rango|nombre|clase1|clase2|nac|est|unidad|pel|esc|rol
    # 0      1      2     3      4   5     6    7   8   9

    print("Creando Miembros...")
    script_dir = os.path.dirname(os.path.realpath(__file__))
    datos_iniciales = script_dir + '\\datos_iniciales.txt'
    with open(datos_iniciales) as txt:
        for line in txt.readlines():
            array = line.replace('\n', '').split(' ')
            # print(array)
            user = User.objects.create_user(username=array[1], email=array[1] + "@zrarmy.com",
                                            password=array[1].lower())
            miembro = Miembro.objects.get(nombre=array[1])
            miembro.nombre = array[1]
            # print(miembro.nombre)
            miembro.rango = Rango.objects.get(abreviatura=array[0])
            # print(miembro.rango)
            miembro.clase1 = Clase.objects.get(abreviatura=array[2])
            # print(miembro.clase1)
            miembro.clase2 = Clase.objects.get(abreviatura=array[3])
            # print(miembro.clase2)
            miembro.nacionalidad = Nacionalidad.objects.get(abreviatura=array[4])
            # print(miembro.nacionalidad)

            if array[5] == 'A':
                el_estado = 'Activo'
            elif array[5] == 'R':
                el_estado = 'Reserva'
            miembro.estado = el_estado
            # print(miembro.estado)
            miembro.unidad = Unidad.objects.get(abreviatura=array[6])
            # print(miembro.unidad)
            miembro.peloton = array[7]
            # print(miembro.peloton)
            miembro.escuadra = array[8]
            # print(miembro.escuadra)
            miembro.rol = Rol.objects.get(abreviatura=array[9])
            # print(miembro.rol)
            miembro.save()
            # print(miembro)


def crea_mision_ofi(fecha, campana, i, miembros):
    mision = Mision()
    mision.nombre = "Mision " + str(i) + " de " + fecha.strftime("%b")
    mision.fecha_finalizada = fecha
    mision.fecha_creacion = fecha
    mision.tipo = Mision.TIPO_CAMPANA
    mision.estado = Mision.ESTADO_FINALIZADA
    mision.campana = campana
    mision.descripcion = "Campaña jugada durante el mes de " + fecha.strftime("%B")
    mision.autor = miembros.order_by('?').first()
    mision.save()

    t = datetime.datetime.strptime("02:30:15", '%H:%M:%S')
    delta = datetime.timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    for m in miembros:
        if random.choice([True, False]):
            a = Asistencia.objects.create(mision=mision, miembro=m, fecha=mision.fecha_finalizada,
                                          asistencia=Asistencia.ASIST_ASISTE, tiempo_de_sesion=delta)
        else:
            if random.choice([True, False]):
                a = Asistencia.objects.create(mision=mision, miembro=m, fecha=mision.fecha_finalizada,
                                              asistencia=Asistencia.ASIST_FALTA, tiempo_de_sesion=delta)
            else:
                a = Asistencia.objects.create(mision=mision, miembro=m, fecha=mision.fecha_finalizada,
                                              asistencia=Asistencia.ASIST_TARDE, tiempo_de_sesion=delta)


def crea_mision_asp(fecha, camp_asp, ai, aspirantes):
    mis_asp = Mision()
    mis_asp.nombre = "Aspirantes curso " + str(ai) + " Camada de " + fecha.strftime("%b")
    mis_asp.fecha_finalizada = fecha
    mis_asp.fecha_creacion = fecha
    mis_asp.tipo = Mision.TIPO_CURSO
    mis_asp.estado = Mision.ESTADO_FINALIZADA
    mis_asp.campana = camp_asp
    mis_asp.descripcion = "Dia " + str(ai) + " de la camada Aspirante del mes de " + fecha.strftime("%B")
    mis_asp.save()
    t = datetime.datetime.strptime("02:30:15", '%H:%M:%S')
    delta = datetime.timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    for m in aspirantes:
        if random.choice([True, False]):
            a = Asistencia.objects.create(mision=mis_asp, miembro=m, fecha=mis_asp.fecha_finalizada,
                                          asistencia=Asistencia.ASIST_ASISTE, tiempo_de_sesion=delta)
        else:
            if random.choice([True, False]):
                a = Asistencia.objects.create(mision=mis_asp, miembro=m, fecha=mis_asp.fecha_finalizada,
                                              asistencia=Asistencia.ASIST_FALTA, tiempo_de_sesion=delta)
            else:
                a = Asistencia.objects.create(mision=mis_asp, miembro=m, fecha=mis_asp.fecha_finalizada,
                                              asistencia=Asistencia.ASIST_TARDE, tiempo_de_sesion=delta)


def crea_mision_impro(fecha, miembros):
    mision = Mision()
    mision.nombre = "Impro del dia " + fecha.strftime("%Y / %d / %b")
    mision.fecha_finalizada = fecha
    mision.fecha_creacion = fecha
    mision.tipo = Mision.TIPO_IMPROVISADA
    mision.estado = Mision.ESTADO_FINALIZADA
    mision.descripcion = "Improvisada"
    mision.autor = miembros.order_by('?').first()
    mision.save()
    t = datetime.datetime.strptime("02:30:15", '%H:%M:%S')
    delta = datetime.timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
    for m in miembros:
        if random.choice([True, False]):
            a = Asistencia.objects.create(mision=mision, miembro=m, fecha=mision.fecha_finalizada,
                                          asistencia=Asistencia.ASIST_ASISTE, tiempo_de_sesion=delta)
        else:
            if random.choice([True, False]):
                a = Asistencia.objects.create(mision=mision, miembro=m, fecha=mision.fecha_finalizada,
                                              asistencia=Asistencia.ASIST_FALTA, tiempo_de_sesion=delta)
            else:
                a = Asistencia.objects.create(mision=mision, miembro=m, fecha=mision.fecha_finalizada,
                                              asistencia=Asistencia.ASIST_TARDE, tiempo_de_sesion=delta)


def generar_misiones():
    print("Generado misiones, campañas y asistencias")
    # Genero campañas
    fecha = datetime.date(2019, 1, 1)
    fecha_fin = datetime.date(now().year, now().month, now().day)
    # fecha_fin = datetime.date(2019, 2, 1)
    dia = datetime.timedelta(days=1)
    mes = 0
    miembros = Miembro.objects.all().exclude(rango__abreviatura='Asp')
    aspirantes = Miembro.objects.filter(rango__abreviatura='Asp')
    while fecha <= fecha_fin:

        if mes < fecha.month:  # cambio de mes y creo campaña
            i = 1  # numero de mision dentro de la campaña
            ai = 1  # numero de curso asp
            mes += 1
            campana = Campana()
            campana.nombre = "Campaña de " + fecha.strftime("%B")
            campana.tipo = Campana.TIPO_CAMPANA
            campana.estado = Campana.ESTADO_FINALIZADA
            campana.save()

            camp_asp = Campana()
            camp_asp.nombre = "Camada Asp. de " + fecha.strftime("%B")
            camp_asp.tipo = Campana.TIPO_CURSO
            camp_asp.estado = Campana.ESTADO_FINALIZADA
            camp_asp.save()

        if fecha.weekday() == 1:  # martes
            crea_mision_ofi(fecha, campana, i, miembros)
            i += 1
            crea_mision_asp(fecha, camp_asp, ai, aspirantes)
            ai += 1
        elif fecha.weekday() == 2:  # miercoles
            crea_mision_asp(fecha, camp_asp, ai, aspirantes)
            ai += 1
        elif fecha.weekday() == 3:  # jueves
            crea_mision_ofi(fecha, campana, i, miembros)
            i += 1
        else:
            if random.choice([True, False]):  # 20% de chance de generar una impro un dia no oficial
                crea_mision_impro(fecha, miembros)

        fecha = fecha + dia  # Avanzo el dia en el calendario
    print("FINALIZADO AGREAGAR MISIONES Y ASISTENCIAS")


def procesa_xsl():
    print("Procesando planillas de asistencia")
    crea_misiones_xls()  # Crea las misiones del año

    from openpyxl import load_workbook
    import datetime

    workbook = load_workbook(filename="Asistencia_ZR_2019.xlsx", data_only=True)
    year = 2019
    mes = 0
    lista_valores = ['A', 'F', 'J', 'L', 'R', 'O', 'T', 'NP', None]
    for sheet in workbook.worksheets:
        if sheet.title == "Abril":
            mes = 4
        elif sheet.title == "Agosto":
            mes = 8
        elif sheet.title == "Diciembre":
            mes = 12
        elif sheet.title == "Enero":
            mes = 1
        elif sheet.title == "Febrero":
            mes = 2
        elif sheet.title == "Julio":
            mes = 7
        elif sheet.title == "Junio":
            mes = 6
        elif sheet.title == "Marzo":
            mes = 3
        elif sheet.title == "Mayo":
            mes = 5
        elif sheet.title == "Noviembre":
            mes = 11
        elif sheet.title == "Octubre":
            mes = 10
        elif sheet.title == "Septiembre":
            mes = 9
        max_row = sheet.max_row
        max_col = sheet.max_column
        for i in range(2, max_row + 1):  # itera rows, empiezo del 2do porque el 1ro tiene los titulos de columna
            for j in range(1, max_col + 1):  # itera columnas
                cell = sheet.cell(row=i, column=j)  # obtiene cell
                valorcell = str(cell.value).strip() if cell.value is not None else cell.value
                if j == 1:
                    x_rango = valorcell
                    x_rango = str(x_rango).strip('.')
                elif j == 2:
                    x_nombre = valorcell
                elif j == 3:
                    x_c1 = valorcell
                elif j == 4:
                    x_c2 = valorcell
                elif j == 5:
                    x_pais = valorcell
                elif j == 6:
                    x_estado = valorcell
                elif j == 7:
                    x_escuadra = valorcell
                elif j == 8:
                    x_rol = valorcell
                if j == 9:
                    crea_miembro_if_not_exists(x_rango, x_nombre, x_c1, x_c2, x_pais, x_estado, x_escuadra, x_rol)

                if j > 8:  # aca comienzan las asistencias
                    day = int(sheet.cell(row=1, column=j).value)
                    fecha = datetime.datetime(day=day, month=mes, year=year)
                    mision = Mision.objects.get(fecha_finalizada=fecha)
                    miembro = Miembro.objects.get(nombre__iexact=x_nombre)

                    x_asistencia = valorcell
                    if x_asistencia is None:
                        x_asistencia = Asistencia.ASIST_FALTA
                    elif x_asistencia == 'A':
                        x_asistencia = Asistencia.ASIST_ASISTE
                    elif x_asistencia == 'R':
                        x_asistencia = Asistencia.ASIST_RESERVA
                    elif x_asistencia == 'T':
                        x_asistencia = Asistencia.ASIST_TARDE
                    elif x_asistencia == 'J':
                        x_asistencia = Asistencia.ASIST_JUSTIFICADA
                    elif x_asistencia == 'L':
                        x_asistencia = Asistencia.ASIST_LICENCIA
                    elif x_asistencia == 'F':
                        x_asistencia = Asistencia.ASIST_FALTA
                    elif x_asistencia == 'O':
                        x_asistencia = Asistencia.ASIST_ASISTE
                    elif x_asistencia == 'NP':
                        x_asistencia = Asistencia.ASIST_FALTA
                    else:
                        x_asistencia = Asistencia.ASIST_FALTA

                    # TIEMPO DE SESION APROXIMADO
                    if x_asistencia == Asistencia.ASIST_ASISTE:
                        delta = datetime.timedelta(hours=2)
                    elif x_asistencia == Asistencia.ASIST_TARDE:
                        delta = datetime.timedelta(hours=1)
                    else:
                        delta = datetime.timedelta(hours=0)

                    a = Asistencia.objects.create(mision=mision, miembro=miembro, fecha=mision.fecha_finalizada,
                                                  asistencia=x_asistencia, tiempo_de_sesion=delta)


def crea_misiones_xls():
    from openpyxl import load_workbook
    import datetime

    workbook = load_workbook(filename="Asistencia_ZR_2019.xlsx", data_only=True)
    year = 2019
    mes = 0
    lista_fechas_mision = []
    for sheet in workbook.worksheets:
        if sheet.title == "Abril":
            mes = 4
        elif sheet.title == "Agosto":
            mes = 8
        elif sheet.title == "Diciembre":
            mes = 12
        elif sheet.title == "Enero":
            mes = 1
        elif sheet.title == "Febrero":
            mes = 2
        elif sheet.title == "Julio":
            mes = 7
        elif sheet.title == "Junio":
            mes = 6
        elif sheet.title == "Marzo":
            mes = 3
        elif sheet.title == "Mayo":
            mes = 5
        elif sheet.title == "Noviembre":
            mes = 11
        elif sheet.title == "Octubre":
            mes = 10
        elif sheet.title == "Septiembre":
            mes = 9

        # Creo una campaña Mensual
        camp = Campana.objects.create(nombre="Camp. " + sheet.title, tipo=Campana.TIPO_CAMPANA,
                                      estado=Campana.ESTADO_FINALIZADA, descripcion=str(mes))

        # RECORRO LOS DIAS DE CADA MES (columnas entre "K" y la primera columna con signo "%")
        max_col = sheet.max_column
        # TODO AJUSTAR ACÁ EL RANGO DONDE ESTAN COLUMNAS DE LOS DIAS
        for col in sheet.iter_cols(min_col=9, max_col=max_col, min_row=1, max_row=1):
            for cell in col:
                day = int(cell.value)
                dt = datetime.date(year, mes, day)
                lista_fechas_mision.append(dt)

    for mision_fecha in lista_fechas_mision:
        mision = Mision()
        mision.nombre = "Mision (" + str(mision_fecha) + ")"
        mision.fecha_finalizada = mision_fecha
        mision.fecha_creacion = mision_fecha
        if mision_fecha.weekday() in [1, 3]:  # SI ES MARTES O JUEVES ES OFICIAL
            mision.tipo = Mision.TIPO_CAMPANA
        else:
            mision.tipo = Mision.TIPO_IMPROVISADA
        mision.estado = Mision.ESTADO_FINALIZADA
        mision.descripcion = "Misión Oficial del día " + str(mision_fecha)
        # Asigno campaña a mision
        mes = mision_fecha.month
        campana = Campana.objects.get(descripcion=mes)
        mision.campana = campana
        mision.save()

def crea_miembro_if_not_exists(rango, nombre, c1, c2, pais, estado, escuadra, rol):
    try:
        miembro = Miembro.objects.get(nombre__iexact=nombre)
    except Miembro.DoesNotExist:
        user = User.objects.create_user(username=nombre, email=nombre + "@zrarmy.com",
                                        password=nombre.lower())
        miembro = Miembro.objects.get(user=user)

    miembro.nombre = nombre
    miembro.rango = Rango.objects.get(abreviatura=rango)
    if c1 is None:
        miembro.clase1 = Clase.objects.get(abreviatura__iexact='FL')
    else:
        miembro.clase1 = Clase.objects.get(abreviatura__iexact=c1)
    if c2 is None:
        miembro.clase2 = Clase.objects.get(abreviatura__iexact='FL')
    else:
        miembro.clase2 = Clase.objects.get(abreviatura__iexact=c2)

    if pais is None:
        pais = "ZW"
    miembro.nacionalidad = Nacionalidad.objects.get(abreviatura=pais.upper())

    if estado == 'A':
        miembro.estado = Miembro.ESTADO_ACTIVO
    else:
        miembro.estado = Miembro.ESTADO_RESERVA
    if escuadra == "1° P - 2°M":
        miembro.unidad = Unidad.objects.get(abreviatura="IMZR")
        miembro.peloton = 2
    elif escuadra == "1° PP - 1°Pa":
        miembro.unidad = Unidad.objects.get(abreviatura="PAR")
        miembro.peloton = 1
    elif escuadra == "ALTM":
        miembro.unidad = Unidad.objects.get(abreviatura="ALTM")
        miembro.peloton = 1
    elif escuadra == "Caballeria":
        miembro.unidad = Unidad.objects.get(abreviatura="CAB")
        miembro.peloton = 1
    elif escuadra == "Espectro":
        miembro.unidad = Unidad.objects.get(abreviatura="ECHO")
        miembro.peloton = 1
    elif escuadra == "FAZR":
        miembro.unidad = Unidad.objects.get(abreviatura="FAZR")
        miembro.peloton = 1
    else:
        miembro.unidad = Unidad.objects.get(abreviatura="IMZR")
        miembro.peloton = 1
    miembro.save()
